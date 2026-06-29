"""Server-side spreadsheet parse endpoint — ``POST /api/parse-file``.

This module is the backend half of the SheetJS → server-side migration. The
frontend formerly parsed ``.xlsx`` in the browser via the ``xlsx`` (SheetJS)
package, which carried two unpatched HIGH advisories (GHSA-4r6h-8v6p-xvw6
prototype pollution, GHSA-5pgg-2g8v-p4x9 ReDoS). That package was removed; the
browser now POSTs the raw file here and receives a typed
:class:`~backend.models.parsed_file.ParsedFileResponse`.

Why openpyxl directly, not ``pandas.read_excel``
------------------------------------------------
``pandas.read_excel`` is a convenience wrapper that silently coerces dtypes,
fills NaN, and infers types before the data is ever inspected — it *fixes*
before it *detects*. That is precisely backwards for a data-quality platform.
We instead drive :func:`openpyxl.load_workbook` and own the cell-iteration loop,
so null cells survive as ``None``, dates serialise losslessly to ISO 8601, and
anything we cannot represent faithfully (an Excel error literal, a formula with
no cached result) is surfaced as a structured parse error rather than mutated.

Two-pass read
-------------
``data_only=True`` returns a formula's *cached result* but erases the fact that
the cell was ever a formula. To still flag a formula whose result is missing
(workbook never recalculated) we read the workbook twice: once with
``data_only=False`` to learn which cells are formulas, once with
``data_only=True`` to read their computed values. Both passes are ``read_only``
to keep memory bounded on large sheets.

Scope: openpyxl reads the OOXML formats (``.xlsx`` / ``.xlsm``) only. The legacy
binary ``.xls`` (BIFF) format is rejected with a 400 — it would need a separate
reader and is handled as an explicit product decision, not a silent fallback.
"""

from __future__ import annotations

from datetime import date, datetime, time
from io import BytesIO
from typing import Any

import structlog
from fastapi import APIRouter, File, HTTPException, UploadFile
from openpyxl import load_workbook
from openpyxl.worksheet._read_only import ReadOnlyWorksheet

from backend.models.parsed_file import CellParseError, ParsedFileResponse

logger = structlog.get_logger(__name__)

router = APIRouter(prefix="/api", tags=["parse"])

# openpyxl reads OOXML only. ``.xls`` (legacy BIFF) is intentionally excluded.
_SUPPORTED_EXTENSIONS = frozenset({"xlsx", "xlsm"})


class SpreadsheetParseError(Exception):
    """Raised when a workbook cannot be read or contains no usable data.

    Carries a user-facing message; the route maps it to an HTTP 400 so the
    frontend can surface it the same way the former client-side parser did.
    """


def _extension(filename: str | None) -> str:
    """Return the lowercased extension of *filename*, or ``""`` if none."""
    if not filename or "." not in filename:
        return ""
    return filename.rsplit(".", 1)[-1].lower()


def _map_formulas(ws: ReadOnlyWorksheet) -> dict[tuple[int, int], str]:
    """Map ``(row, column)`` → formula string for every formula cell.

    Read from the ``data_only=False`` pass, where a formula cell's
    ``data_type`` is ``"f"`` and its value is the formula text. Used to flag
    formulas whose cached result is absent in the value pass.
    """
    formulas: dict[tuple[int, int], str] = {}
    for row in ws.iter_rows():
        for cell in row:
            if cell.value is not None and cell.data_type == "f":
                formulas[(cell.row, cell.column)] = str(cell.value)
    return formulas


def _row_number(cells: Any, fallback: int) -> int:
    """Return the 1-based sheet row for a row of cells.

    In read-only mode, blank padding cells are shared ``EmptyCell`` sentinels
    with no coordinates, so we read the row index off the first real cell and
    fall back to a positional estimate if the whole row is empty.
    """
    for cell in cells:
        row = getattr(cell, "row", None)
        if row is not None:
            return row
    return fallback


def _cell_value(
    cell: Any,
    formula_map: dict[tuple[int, int], str],
    column: str,
    row_number: int,
    errors: list[CellParseError],
) -> Any:
    """Convert one openpyxl cell to a JSON-safe value, recording any issue.

    Preserves nulls as ``None`` (never ``""``), serialises temporal values to
    ISO 8601, emits ``None`` for Excel error literals and missing-result
    formulas while appending a :class:`CellParseError` for each. Numbers,
    strings, and booleans pass through unchanged.
    """
    if cell is None:
        return None

    # Excel error literal (#REF!, #N/A, #DIV/0!, ...). Surface, do not coerce.
    if getattr(cell, "data_type", None) == "e":
        errors.append(
            CellParseError(
                row=row_number, column=column, issue="error_cell", raw_value=str(cell.value)
            )
        )
        return None

    value = cell.value

    if value is None:
        # Empty cell, or a formula whose cached result was never computed. The
        # former is valid data (preserved as null); only the latter is flagged.
        # Read-only blank cells are coordinateless EmptyCells — they can never
        # be formulas, so a missing coordinate simply means "ordinary null".
        cell_row = getattr(cell, "row", None)
        cell_col = getattr(cell, "column", None)
        formula = (
            formula_map.get((cell_row, cell_col))
            if cell_row is not None and cell_col is not None
            else None
        )
        if formula is not None:
            errors.append(
                CellParseError(
                    row=row_number, column=column, issue="formula_cell", raw_value=formula
                )
            )
        return None

    # Temporal cells arrive as Python datetime/date/time; ISO 8601 round-trips
    # losslessly and is unambiguous across the wire.
    if isinstance(value, (datetime, date, time)):
        return value.isoformat()

    return value


def parse_spreadsheet(content: bytes, file_type: str) -> ParsedFileResponse:
    """Parse OOXML spreadsheet *content* into a :class:`ParsedFileResponse`.

    Pure function (no HTTP) so it is unit-testable in isolation. The first row
    is treated as the header row, matching the prior client-side behaviour;
    blank header cells become ``"Column N"``. Raises
    :class:`SpreadsheetParseError` on an unreadable or empty workbook.
    """
    try:
        wb_formulas = load_workbook(BytesIO(content), data_only=False, read_only=True)
        wb_values = load_workbook(BytesIO(content), data_only=True, read_only=True)
    except Exception as exc:  # openpyxl raises a variety of types on bad input
        raise SpreadsheetParseError(
            f"Could not read spreadsheet: {exc}"
        ) from exc

    try:
        if not wb_formulas.worksheets or not wb_values.worksheets:
            raise SpreadsheetParseError("Spreadsheet contains no worksheets.")

        formula_map = _map_formulas(wb_formulas.worksheets[0])
        ws = wb_values.worksheets[0]

        rows_iter = ws.iter_rows()
        header_cells = next(rows_iter, None)
        if header_cells is None:
            raise SpreadsheetParseError("Spreadsheet appears to be empty.")

        headers: list[str] = []
        for idx, cell in enumerate(header_cells):
            raw = cell.value
            headers.append(str(raw) if raw not in (None, "") else f"Column {idx + 1}")

        header_row = _row_number(header_cells, 1)

        errors: list[CellParseError] = []
        rows: list[list[Any]] = []
        for offset, row_cells in enumerate(rows_iter):
            row_number = _row_number(row_cells, header_row + 1 + offset)
            row_values = [
                _cell_value(
                    row_cells[i] if i < len(row_cells) else None,
                    formula_map,
                    headers[i],
                    row_number,
                    errors,
                )
                for i in range(len(headers))
            ]
            rows.append(row_values)

        return ParsedFileResponse(
            headers=headers,
            rows=rows,
            total_rows=len(rows),
            file_type=file_type,
            parse_errors=errors,
        )
    finally:
        wb_formulas.close()
        wb_values.close()


@router.post("/parse-file", response_model=ParsedFileResponse)
async def parse_file(file: UploadFile = File(...)) -> ParsedFileResponse:
    """Parse an uploaded ``.xlsx`` / ``.xlsm`` file into structured rows.

    Returns 400 for unsupported extensions (including legacy ``.xls``) and for
    workbooks that cannot be read.
    """
    file_type = _extension(file.filename)
    if file_type not in _SUPPORTED_EXTENSIONS:
        detail = (
            f"Unsupported file type for server-side parse: "
            f"'{file_type or 'unknown'}'. This endpoint handles .xlsx and .xlsm; "
            f"legacy .xls is not supported."
        )
        raise HTTPException(status_code=400, detail=detail)

    content = await file.read()
    try:
        result = parse_spreadsheet(content, file_type)
    except SpreadsheetParseError as exc:
        logger.warning("parse_file.rejected", filename=file.filename, reason=str(exc))
        raise HTTPException(status_code=400, detail=str(exc)) from exc

    logger.info(
        "parse_file.parsed",
        filename=file.filename,
        rows=result.total_rows,
        columns=len(result.headers),
        parse_errors=len(result.parse_errors),
    )
    return result
