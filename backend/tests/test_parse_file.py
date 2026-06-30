"""Tests for the server-side spreadsheet parser (SheetJS migration).

Covers the pure :func:`parse_spreadsheet` function and the HTTP route. Fixtures
build ``.xlsx`` bytes in memory with openpyxl — no test files on disk.

A note on formula cells: openpyxl writes formulas but never evaluates them, so
a workbook produced here has no cached result. That deterministically exercises
the "formula result missing → parse_errors entry" branch, which is exactly the
contract the frontend relies on.
"""

from __future__ import annotations

from datetime import datetime
from io import BytesIO
from typing import Any

import pytest
from fastapi.testclient import TestClient
from openpyxl import Workbook

from backend.api.app import app
from backend.api.parse_file import SpreadsheetParseError, parse_spreadsheet


def _xlsx_bytes(rows: list[list[Any]]) -> bytes:
    """Serialise *rows* (first row = headers) into ``.xlsx`` bytes."""
    wb = Workbook()
    ws = wb.active
    for row in rows:
        ws.append(row)
    buf = BytesIO()
    wb.save(buf)
    return buf.getvalue()


@pytest.fixture
def client() -> TestClient:
    return TestClient(app)


def test_happy_path_headers_and_row_count() -> None:
    content = _xlsx_bytes(
        [
            ["region", "units", "revenue"],
            ["north", 10, 1000.5],
            ["south", 20, 2000.0],
            ["east", 30, 3000.0],
            ["west", 40, 4000.0],
            ["central", 50, 5000.0],
        ]
    )

    result = parse_spreadsheet(content, "xlsx")

    assert result.headers == ["region", "units", "revenue"]
    assert result.total_rows == 5
    assert result.rows[0] == ["north", 10, 1000.5]
    assert result.parse_errors == []


def test_null_cells_preserved_as_none_not_empty_string() -> None:
    content = _xlsx_bytes(
        [
            ["region", "revenue"],
            ["north", None],
            [None, 2000.0],
        ]
    )

    result = parse_spreadsheet(content, "xlsx")

    # Nulls survive as None — never coerced to "" — and are NOT parse errors.
    assert result.rows[0] == ["north", None]
    assert result.rows[1] == [None, 2000.0]
    assert result.parse_errors == []


def test_date_cells_serialise_to_iso_8601() -> None:
    content = _xlsx_bytes(
        [
            ["event", "when"],
            ["launch", datetime(2026, 6, 26, 9, 30, 0)],
        ]
    )

    result = parse_spreadsheet(content, "xlsx")

    assert result.rows[0][0] == "launch"
    assert result.rows[0][1] == "2026-06-26T09:30:00"


def test_formula_without_cached_result_is_recorded_not_dropped() -> None:
    content = _xlsx_bytes(
        [
            ["a", "b", "total"],
            [1, 2, "=A2+B2"],
        ]
    )

    result = parse_spreadsheet(content, "xlsx")

    # openpyxl never computed the formula, so the value is absent: we emit None
    # for the cell AND surface a structured parse error rather than silently
    # dropping it.
    assert result.rows[0][2] is None
    assert len(result.parse_errors) == 1
    err = result.parse_errors[0]
    assert err.issue == "formula_cell"
    assert err.column == "total"
    assert "A2+B2" in err.raw_value


def test_excel_error_literal_is_recorded_and_nulled() -> None:
    wb = Workbook()
    ws = wb.active
    ws.append(["value"])
    ws["A2"] = "#REF!"  # openpyxl binds known error codes to data_type "e"
    buf = BytesIO()
    wb.save(buf)

    result = parse_spreadsheet(buf.getvalue(), "xlsx")

    assert result.rows[0][0] is None
    assert len(result.parse_errors) == 1
    assert result.parse_errors[0].issue == "error_cell"
    assert result.parse_errors[0].raw_value == "#REF!"


def test_blank_header_cell_becomes_column_n() -> None:
    content = _xlsx_bytes(
        [
            ["region", None, "revenue"],
            ["north", "x", 1000],
        ]
    )

    result = parse_spreadsheet(content, "xlsx")

    assert result.headers == ["region", "Column 2", "revenue"]


def test_header_only_sheet_yields_zero_rows() -> None:
    content = _xlsx_bytes([["region", "revenue"]])

    result = parse_spreadsheet(content, "xlsx")

    assert result.headers == ["region", "revenue"]
    assert result.total_rows == 0
    assert result.rows == []


def test_empty_workbook_raises() -> None:
    wb = Workbook()
    buf = BytesIO()
    wb.save(buf)

    with pytest.raises(SpreadsheetParseError):
        parse_spreadsheet(buf.getvalue(), "xlsx")


def test_unreadable_content_raises() -> None:
    with pytest.raises(SpreadsheetParseError):
        parse_spreadsheet(b"this is not a spreadsheet", "xlsx")


def test_route_happy_path(client: TestClient) -> None:
    content = _xlsx_bytes([["region", "revenue"], ["north", 1000]])

    response = client.post(
        "/api/parse-file",
        files={"file": ("data.xlsx", content, "application/vnd.openxmlformats")},
    )

    assert response.status_code == 200
    body = response.json()
    assert body["headers"] == ["region", "revenue"]
    assert body["total_rows"] == 1
    assert body["parse_errors"] == []


def test_route_rejects_legacy_xls(client: TestClient) -> None:
    response = client.post(
        "/api/parse-file",
        files={"file": ("legacy.xls", b"whatever", "application/vnd.ms-excel")},
    )

    assert response.status_code == 400
    assert ".xls" in response.json()["detail"]


def test_route_rejects_unsupported_extension(client: TestClient) -> None:
    response = client.post(
        "/api/parse-file",
        files={"file": ("deck.pptx", b"whatever", "application/octet-stream")},
    )

    assert response.status_code == 400
    assert "Unsupported file type" in response.json()["detail"]
